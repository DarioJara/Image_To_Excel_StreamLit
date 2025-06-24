import streamlit as st
import pytesseract
from PIL import Image
import openpyxl
from io import BytesIO
import platform

# Configura ruta de tesseract si es Windows
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def convertir_imagen_a_excel(imagen):
    # Cargar imagen
    img = Image.open(imagen)

    # OCR con configuración para solo números
    custom_config = r'--oem 3 --psm 6 outputbase digits'
    data = pytesseract.image_to_string(img, config=custom_config)

    rows = data.strip().split('\n')
    table = [row.split() for row in rows if row.strip() != '']

    # Crear Excel en memoria
    wb = openpyxl.Workbook()
    ws = wb.active

    for i, row in enumerate(table, start=1):
        for j, val in enumerate(row, start=1):
            try:
                ws.cell(row=i, column=j, value=float(val))
            except ValueError:
                ws.cell(row=i, column=j, value=val)

    # Guardar archivo Excel en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Streamlit UI
st.title("Convertidor de imagen a Excel")
st.write("Sube una imagen que contenga números (como una tabla escaneada) y la convertiré en un archivo Excel.")

uploaded_file = st.file_uploader("Selecciona una imagen", type=["png", "jpg", "jpeg", "bmp", "tiff"])

if uploaded_file is not None:
    st.image(uploaded_file, caption="Imagen subida", use_column_width=True)

    if st.button("Convertir a Excel"):
        with st.spinner("Procesando imagen..."):
            excel_data = convertir_imagen_a_excel(uploaded_file)

        st.success("Conversión completada.")
        st.download_button(
            label="Descargar Excel",
            data=excel_data,
            file_name="resultado_dariojara.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.markdown("---")
st.markdown(
    "<div style='text-align: center; font-size: 14px; color: gray;'>Elaborado por Dario Jara</div>",
    unsafe_allow_html=True
)